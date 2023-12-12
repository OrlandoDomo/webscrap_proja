import pandas as pd

from io import BytesIO
from openpyxl import load_workbook
from SharePoint_API import SharePoint

from inei_data import download_inei_data_pbi
from midagri_data import download_midagri_data


FILE_NAME= 'pbi_valor_copia.xlsx'
FOLDER_NAME= '01. Bases de datos/Agropecuario/1. Proyecciones de PBI Agro/1. Proyecciones'

def upload_SharePoint_file(fn):
    content = fn()
    SharePoint().upload_file(FILE_NAME, FOLDER_NAME, content)

def update_monthly_midagri_data_sheet(fn, ):
    midagri_excel= download_midagri_data() 

    df=pd.read_excel(midagri_excel, sheet_name='vbp_publ')

    df= df.drop(labels=['Unnamed: 8' ,'Unnamed: 9'],axis=1)
    df= df.drop([0,1,2,3,99,100,101])
    df1=df.iloc[:,:8].reset_index(drop=True)

    for i in range(len(df1)):
        value= df1.iloc[i].iloc[1]
        if pd.isna(value):
            df1.iloc[i].iloc[1]=df1.iloc[i].iloc[0]

    df1= df1.drop(labels='VALOR DE LA PRODUCCIÓN (VBP) AGROPECUARIA',axis=1)

    for i in range(len(df1.iloc[0])):
        value=df1.iloc[0].iloc[i]
        if pd.isna(value):
            value=temp
        else:
            temp=value
        df1.iloc[0].iloc[i]= temp
        df1.rename(columns={ df1.columns[i]: "{:s} {:s}".format(df1.iloc[0].iloc[i], str(df1.iloc[1].iloc[i]).split(' ')[0]) }, inplace = True)

    df1= df1.drop([0,1]).reset_index(drop=True)

    ## Es necesario cambiar los siguientes nombres: Cafe, Cana de azucar, Maiz amilaceo, Arandanos, Arverja, Huevos y Leche fresca.
    ## Seria necesario cambiar "total"es a "Subsector agricola"/"Subsector pecuario"

    products= [p.lower() for p in df1['Principales productos nan'].tolist()]

    i=0
    for p in df1[df1.columns[0]]:
        df1[df1.columns[0]].iloc[i]= p.lower()
        i=i+1    

    midagri_file_obj = SharePoint().download_file(FILE_NAME, FOLDER_NAME)
    midagri_data = BytesIO(midagri_file_obj)
    wb = load_workbook(midagri_data)
    ws = wb['MM soles 2007 (1)']

    highlight = PatternFill(start_color='00FF6600',
                   end_color='00FF6600',
                   fill_type='solid')

    header_row=[]

    # Esto modifical el excel

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=3, max_col= 3):
        for cell in row:
            if cell.internal_value=='Periodo':
                header_row.append(cell.row)
            if cell.internal_value==int(year+month):
                target_row= cell.row

    for col in ws.iter_cols(min_row=target_row, max_row=target_row, min_col=2, max_col= ws.max_column - 5):
        for cell in col:
            product= str(ws[cell.column_letter+f'{header_row[0]}'].value)
            if product.lower() in products:
                i= df1[ df1[df1.columns[0]] == product.lower()].index
                cell.value= df1.iloc[i,5].values[0]
                cell.fill= highlight 
            if product.lower()=='total':
                subsector_cell= ws[cell.column_letter+f'{header_row[0]-1}']
                subsector= ws[parent_of_merged_cell(subsector_cell)].value
                i= df1[ df1[df1.columns[0]] == 'subsector '+ subsector.lower()].index
                cell.value= df1.iloc[i,5].values[0]
                cell.fill= highlight 

    buffer = BytesIO()
    wb.save(buffer)
    #wb.save(path_to_excel)




def parent_of_merged_cell(cell):
    """ Find the parent of the merged cell by iterating through the range of merged cells """
    sheet = cell.parent
    child_coord = cell.coordinate

    # Note: if there are many merged cells in a large spreadsheet, this may become inefficient
    for merged in sheet.merged_cells.ranges:
        if child_coord in merged:
            return merged.start_cell.coordinate
    return None

def cell_value(cell):
    """ Reads the value of a cell, if cell is within a merged cell,
        find the first cell in the merged cell and get its value
    """
    if isinstance(cell, openpyxl.cell.cell.Cell):
        return cell.value
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        coord = parent_of_merged_cell(cell)
        parent = cell.parent[coord]
        return parent.value